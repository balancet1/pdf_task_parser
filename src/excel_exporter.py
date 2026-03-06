import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

class ExcelExporter:
    def __init__(self, df):
        """
        Инициализация экспортера с DataFrame
        """
        self.df = df
        self.wb = None
        self.ws = None
    
    def export(self, filename="output/tasks.xlsx", sheet_name="Задачи"):
        """
        Экспорт DataFrame в Excel с форматированием
        """
        # Создаем директорию, если её нет
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        # Создаем Excel файл с форматированием через openpyxl
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name
        
        # Записываем данные
        self._write_data()
        
        # Применяем форматирование
        self._apply_formatting()
        
        # Сохраняем файл
        self.wb.save(filename)
        print(f"✅ Excel файл сохранен: {filename}")
        return filename
    
    def _write_data(self):
        """Запись данных из DataFrame в лист"""
        # Записываем заголовки
        headers = list(self.df.columns)
        for col_idx, header in enumerate(headers, 1):
            self.ws.cell(row=1, column=col_idx, value=header)
        
        # Записываем данные
        for row_idx, row in self.df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=row_idx + 2, column=col_idx, value=value)
                
                # Если это дата - форматируем
                if isinstance(value, (datetime, pd.Timestamp)):
                    cell.number_format = 'DD.MM.YYYY'
    
    def _apply_formatting(self):
        """Применяет красивое форматирование к таблице"""
        
        # Стили
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        date_alignment = Alignment(horizontal='center', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Форматируем заголовки
        for col in range(1, len(self.df.columns) + 1):
            cell = self.ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Форматируем данные
        for row in range(2, self.ws.max_row + 1):
            for col in range(1, len(self.df.columns) + 1):
                cell = self.ws.cell(row=row, column=col)
                cell.border = border
                
                # Для колонки с датой - центрирование
                if self.df.columns[col-1] in ['Срок', 'Дата']:
                    cell.alignment = date_alignment
                else:
                    cell.alignment = cell_alignment
        
        # Автоподбор ширины колонок
        for col in self.ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Ограничиваем максимальную ширину
            adjusted_width = min(max_length + 2, 80)
            self.ws.column_dimensions[col_letter].width = adjusted_width
        
        # Фиксируем заголовок
        self.ws.freeze_panes = 'A2'
    
    def export_multiple_sheets(self, filename="output/tasks_report.xlsx", extra_data=None):
        """
        Экспорт в несколько листов
        """
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Основной лист с задачами
            self.df.to_excel(writer, sheet_name='Все задачи', index=False)
            
            # Статистика по ответственным
            if 'Ответственный' in self.df.columns:
                stats = self.df['Ответственный'].value_counts().reset_index()
                stats.columns = ['Ответственный', 'Количество задач']
                stats.to_excel(writer, sheet_name='Статистика', index=False)
            
            # Задачи по датам (если есть колонка с датой)
            if 'Срок' in self.df.columns:
                # Простая сортировка
                self.df.sort_values('Срок').to_excel(
                    writer, sheet_name='По срокам', index=False
                )
            
            # Дополнительные данные, если переданы
            if extra_data:
                for sheet_name, data in extra_data.items():
                    if isinstance(data, pd.DataFrame):
                        data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"✅ Многостраничный отчет сохранен: {filename}")
